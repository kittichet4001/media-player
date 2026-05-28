
from openpyxl import load_workbook
input_file = ".\\Thai\\Youtube_Chart.xlsx"
wb = load_workbook(input_file)
sheet = wb.active
row_cnt = 0
output_file = ".\\Thai\\Youtube_top_thaisong.txt"
with open(output_file, 'w', encoding='utf-8') as f:
    for row in sheet.iter_rows(values_only=True):
        row_cnt += 1
        if row_cnt == 1: # ข้ามแถวแรก (หัวตาราง)
            continue
        last_element = row[-1] # เลือกตัวสุดท้ายของแถว
        if last_element is not None:
            f.write(f"{last_element}\n")